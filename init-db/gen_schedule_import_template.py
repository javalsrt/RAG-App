# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '课表导入模板'

# 说明行
ws.merge_cells('A1:L1')
title = ws['A1']
title.value = '课表导入模板（请按示例格式填写，支持 xlsx/docx/pdf/图片）'
title.font = Font(bold=True, size=14, color='FFFFFF')
title.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
title.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 32

# 表头
headers = ['序号', '课程名称', '授课教师', '星期', '节次', '开始时间', '结束时间', '教室', '周次', '总课时', '班级', '学期']
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

for col, h in enumerate(headers, 1):
    c = ws.cell(row=2, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin

# 示例数据：每周 2 节课，连堂示例
# credit 按每周实际占用节数填写，用于排课弹窗的周课时上限
examples = [
    [1, 'Python编程入门', '张明', '周二', '1-1节', '08:10', '08:50', '实训楼A202', '[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16]', 2, '人工智能2024级1班', '2025-2026-2'],
    [2, 'Python编程入门', '张明', '周四', '4-4节', '10:40', '11:20', '实训楼A202', '[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16]', 2, '人工智能2024级1班', '2025-2026-2'],
    [3, '深度学习实践', '李华', '周一', '5-6节', '15:10', '16:40', '实训楼B301', '[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16]', 2, '人工智能2024级1班', '2025-2026-2'],
]

for row_idx, row_data in enumerate(examples, 3):
    for col, val in enumerate(row_data, 1):
        c = ws.cell(row=row_idx, column=col, value=val)
        c.alignment = center
        c.border = thin

# 说明 sheet
ws2 = wb.create_sheet('填写说明')
notes = [
    ['字段', '说明'],
    ['课程名称', '必须与系统中的 course 表课程名一致，或新名称会自动创建课程'],
    ['授课教师', '管理员导入时必填，需与 teacher 表真实姓名一致'],
    ['星期', '1=周一，2=周二，...，6=周六，周日不排课'],
    ['节次', '格式：开始节-结束节，如 1-1 表示第1节，5-6 表示第5、6节连堂'],
    ['开始时间/结束时间', '必须严格匹配系统 8 节课表时间段'],
    ['教室', '教室名称，如 实训楼A202'],
    ['周次', '必须完整展开为 JSON 数组，如 [1,2,3,4,5,6]，禁止写 [1-6]'],
    ['总课时', '在当前实现中作为每周排课上限，建议填每周实际占用节数'],
    ['班级', '完整班级名或简写，如 人工智能2024级1班 / 人工智能2401班'],
    ['学期', '如 2025-2026-2 或 2026-暑假培训'],
    ['', ''],
    ['系统 8 节课表时间段', ''],
    ['第 1 节', '08:10 - 08:50'],
    ['第 2 节', '09:00 - 09:40'],
    ['第 3 节', '09:50 - 10:30'],
    ['第 4 节', '10:40 - 11:20'],
    ['第 5 节', '15:10 - 15:50'],
    ['第 6 节', '16:00 - 16:40'],
    ['第 7 节', '19:50 - 20:10'],
    ['第 8 节', '20:20 - 21:00'],
]

for row_idx, (k, v) in enumerate(notes, 1):
    c1 = ws2.cell(row=row_idx, column=1, value=k)
    c2 = ws2.cell(row=row_idx, column=2, value=v)
    if row_idx == 1 or row_idx == 13:
        c1.font = Font(bold=True, color='FFFFFF')
        c2.font = Font(bold=True, color='FFFFFF')
        c1.fill = header_fill
        c2.fill = header_fill
    c1.border = thin
    c2.border = thin
    c1.alignment = center
    c2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 70

widths = [6, 20, 12, 10, 12, 12, 12, 16, 30, 10, 22, 18]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = 'A3'

import os
output_dir = r'C:\Users\jay\Desktop\myBishe\aiStudy\web-admin-react\public\template'
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, 'schedule-import-template.xlsx')
wb.save(output_path)
print(f'已生成：{output_path}')
