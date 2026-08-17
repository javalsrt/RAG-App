# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import random

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '学生导入数据'

# 样式
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center')

# 表头
headers = ['学号', '姓名', '用户名', '班级名称', '专业', '年级', '邮箱', '手机号']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin

# 姓名生成池
surnames = list('赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜')
names_single = list('伟芳娜敏静丽强磊军洋勇艳杰涛明超秀兰霞平刚桂英健峰颖婷宇翔飞鑫淼')
names_double = ['子涵', '梓涵', '子轩', '梓睿', '雨涵', '若曦', '浩然', '俊杰', '思琪', '梦瑶',
                '嘉怡', '雅婷', '志强', '文博', '宇辰', '欣怡', '可馨', '紫萱', '昊天', '晨阳']

# 生成30名学生数据
students = []
used_names = set()
for i in range(1, 31):
    # 学号：2024 + 两位序号
    student_no = f'2024{i:02d}'
    # 用户名：stu + 学号后4位
    username = f'stu{student_no}'
    # 姓名：避免重复
    while True:
        surname = random.choice(surnames)
        if random.random() > 0.5:
            name = surname + random.choice(names_single)
        else:
            name = surname + random.choice(names_double)
        if name not in used_names:
            used_names.add(name)
            break
    # 邮箱
    email = f'{username}@example.com'
    # 手机号
    phone = f'139{random.randint(10000000, 99999999)}'

    students.append((student_no, name, username, '人工智能2024级1班', '人工智能技术', '2024', email, phone))

# 写入数据
for row_idx, row_data in enumerate(students, 2):
    for col_idx, val in enumerate(row_data, 1):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.alignment = center
        c.border = thin

# 列宽
widths = [12, 10, 14, 20, 14, 8, 22, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 冻结表头
ws.freeze_panes = 'A2'

wb.save(r'C:\Users\jay\Desktop\myBishe\aiStudy\init-db\import-class-ai30.xlsx')
print(f'Excel saved with {len(students)} students, class: 人工智能2024级1班')
print('First 5 students:')
for s in students[:5]:
    print(f'  {s[0]} {s[1]} ({s[2]})')
print('Last 5 students:')
for s in students[-5:]:
    print(f'  {s[0]} {s[1]} ({s[2]})')
