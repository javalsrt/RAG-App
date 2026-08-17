# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

# 计算暑假期间的周次和星期
start_date = date(2026, 7, 25)  # 周六
end_date = date(2026, 8, 31)    # 周一

# 计算学期中的周次（假设以7月25日所在周为第1周）
# 找到7月25日所在周的周一
monday_before = start_date - timedelta(days=start_date.weekday())
# 计算总周数
total_days = (end_date - monday_before).days + 1
total_weeks = (total_days + 6) // 7

print(f"暑假起始日(7/25): {start_date.strftime('%Y-%m-%d')} 星期{['一','二','三','四','五','六','日'][start_date.weekday()]}")
print(f"暑假结束日(8/31): {end_date.strftime('%Y-%m-%d')} 星期{['一','二','三','四','五','六','日'][end_date.weekday()]}")
print(f"起始周一: {monday_before.strftime('%Y-%m-%d')}")
print(f"总周数: {total_weeks}")

# 生成周次数组（1到total_weeks）
all_weeks = list(range(1, total_weeks + 1))
print(f"周次范围: {all_weeks}")

# 暑假培训课程设计
# 班级：人工智能2024级1班
# 课程：5门，集中在工作日白天
courses = [
    # (课程名, 教师, 星期几(1-7), 开始节次, 节数, 开始时间, 结束时间, 教室, 周次, 总课时)
    ("Python编程入门", "李明远", 1, 1, 4, "08:00", "11:40", "实训楼A201", [1,2,3,4,5,6], 32),
    ("Python编程入门", "李明远", 3, 1, 4, "08:00", "11:40", "实训楼A201", [1,2,3,4,5,6], 32),
    ("Python编程入门", "李明远", 5, 1, 4, "08:00", "11:40", "实训楼A201", [1,2,3,4,5,6], 32),

    ("机器学习基础", "王思远", 2, 1, 4, "08:00", "11:40", "实训楼A202", [1,2,3,4,5,6], 32),
    ("机器学习基础", "王思远", 4, 1, 4, "08:00", "11:40", "实训楼A202", [1,2,3,4,5,6], 32),
    ("机器学习基础", "王思远", 6, 1, 4, "08:00", "11:40", "实训楼A202", [1,2,3,4,5,6], 32),

    ("深度学习实践", "赵晓峰", 1, 5, 4, "14:00", "17:40", "实训楼B301", [2,3,4,5,6], 32),
    ("深度学习实践", "赵晓峰", 3, 5, 4, "14:00", "17:40", "实训楼B301", [2,3,4,5,6], 32),
    ("深度学习实践", "赵晓峰", 5, 5, 4, "14:00", "17:40", "实训楼B301", [2,3,4,5,6], 32),

    ("数据可视化", "陈雨桐", 2, 5, 4, "14:00", "17:40", "实训楼B302", [2,3,4,5,6], 32),
    ("数据可视化", "陈雨桐", 4, 5, 4, "14:00", "17:40", "实训楼B302", [2,3,4,5,6], 32),
    ("数据可视化", "陈雨桐", 6, 5, 4, "14:00", "17:40", "实训楼B302", [2,3,4,5,6], 32),

    ("自然语言处理入门", "孙文博", 2, 9, 2, "19:00", "20:40", "线上直播", [3,4,5,6], 16),
    ("自然语言处理入门", "孙文博", 4, 9, 2, "19:00", "20:40", "线上直播", [3,4,5,6], 16),
]

print(f"\n课程安排: {len(courses)} 条记录")

# 生成"课表视图"Excel文件（标准课表表格形式，方便AI识别）
wb = openpyxl.Workbook()

# ============== Sheet1: 课表总表（按星期排列的表格） ==============
ws = wb.active
ws.title = '暑假培训课表'

# 标题
ws.merge_cells('A1:H1')
title_cell = ws['A1']
title_cell.value = '人工智能2024级1班 暑假培训课表'
title_cell.font = Font(bold=True, size=16, color='FFFFFF')
title_cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 36

# 副标题
ws.merge_cells('A2:H2')
sub_cell = ws['A2']
sub_cell.value = '培训时间：2026年7月25日 - 2026年8月31日（共6周）'
sub_cell.font = Font(size=11, color='666666')
sub_cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
sub_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 24

# 节次对应时间
time_slots = [
    ('1-2节', '08:00-09:40'),
    ('3-4节', '10:00-11:40'),
    ('5-6节', '14:00-15:40'),
    ('7-8节', '16:00-17:40'),
    ('9-10节', '19:00-20:40'),
]

day_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']

# 表头行（第3行）
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
thin = Border(
    left=Side(style='thin', color='B4B4B4'),
    right=Side(style='thin', color='B4B4B4'),
    top=Side(style='thin', color='B4B4B4'),
    bottom=Side(style='thin', color='B4B4B4')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.cell(row=3, column=1, value='时间')
ws.cell(row=3, column=2, value='节次')
for i, day in enumerate(day_names):
    ws.cell(row=3, column=3 + i, value=day)

for col in range(1, 10):
    c = ws.cell(row=3, column=col)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin

# 构建课表数据结构：按(星期几, 时间段)索引
# 时间段索引：0=1-2节 1=3-4节 2=5-6节 3=7-8节 4=9-10节
schedule_grid = [[None]*7 for _ in range(5)]  # 5个时间段 x 7天

course_colors = {
    'Python编程入门': 'FFF2CC',   # 浅黄
    '机器学习基础': 'E2EFDA',     # 浅绿
    '深度学习实践': 'DDEBF7',     # 浅蓝
    '数据可视化': 'FCE4D6',       # 浅橙
    '自然语言处理入门': 'E4DFEC', # 浅紫
}

for (cname, teacher, dow, start_node, step, st, et, room, weeks, credit) in courses:
    # 确定时间段索引
    if start_node == 1 and step == 4:
        slot_idx = 0  # 占1-4节，合并显示
        schedule_grid[slot_idx][dow-1] = (cname, teacher, room, weeks, 2)  # 2行高
        schedule_grid[1][dow-1] = None  # 清空下一行
    elif start_node == 5 and step == 4:
        slot_idx = 2  # 占5-8节
        schedule_grid[slot_idx][dow-1] = (cname, teacher, room, weeks, 2)
        schedule_grid[3][dow-1] = None
    elif start_node == 9 and step == 2:
        slot_idx = 4
        schedule_grid[slot_idx][dow-1] = (cname, teacher, room, weeks, 1)
    else:
        # 其他情况默认放在第一行
        slot_idx = 0
        if schedule_grid[slot_idx][dow-1] is None:
            schedule_grid[slot_idx][dow-1] = (cname, teacher, room, weeks, 1)

# 填充课表数据
row_num = 4
slot_idx = 0
i = 0
while i < len(time_slots):
    slot_name, slot_time = time_slots[i]
    
    ws.cell(row=row_num, column=1, value=slot_time)
    ws.cell(row=row_num, column=2, value=slot_name)
    
    # 检查这个时间段是否有课程跨多行
    for day_idx in range(7):
        cell_info = schedule_grid[i][day_idx] if i < 5 else None
        if cell_info:
            cname, teacher, room, weeks, row_span = cell_info
            display = f'{cname}\n@{teacher}\n{room}\n第{min(weeks)}-{max(weeks)}周'
            cell = ws.cell(row=row_num, column=3 + day_idx, value=display)
            cell.fill = PatternFill(start_color=course_colors.get(cname, 'FFFFFF'), end_color=course_colors.get(cname, 'FFFFFF'), fill_type='solid')
            cell.font = Font(size=10, bold=True)
            cell.alignment = center
            cell.border = thin
            # 如果跨行，合并单元格
            if row_span > 1:
                ws.merge_cells(start_row=row_num, start_column=3+day_idx, end_row=row_num+row_span-1, end_column=3+day_idx)
    
    # 设置边框
    for col in range(1, 10):
        c = ws.cell(row=row_num, column=col)
        c.border = thin
        c.alignment = center
    
    ws.row_dimensions[row_num].height = 60
    
    # 判断下一行是否被合并占用
    # 简化：每2个节次一行，每次+2节次索引
    # 1-2节 + 3-4节 = 上午2行
    # 但上面逻辑是4节合并，所以跳过下一行
    # 检查是否有课程跨2行
    has_span = False
    for day_idx in range(7):
        info = schedule_grid[i][day_idx] if i < 5 else None
        if info and info[4] > 1:
            has_span = True
            break
    
    if has_span and i + 1 < 5:
        # 已经合并，跳到下下个时间段
        i += 2
        row_num += 1
        # 补充边框
        for col in range(1, 10):
            c = ws.cell(row=row_num, column=col)
            c.border = thin
        ws.cell(row=row_num, column=1, value='').border = thin
        ws.cell(row=row_num, column=2, value='').border = thin
        row_num += 1
    else:
        i += 1
        row_num += 1

# 列宽
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 10
for i in range(7):
    ws.column_dimensions[get_column_letter(3 + i)].width = 18

# ============== Sheet2: 课程明细表（逐行列表，更清晰） ==============
ws2 = wb.create_sheet('课程明细')

headers2 = ['序号', '课程名称', '授课教师', '星期', '节次', '开始时间', '结束时间', '教室', '周次', '总课时', '班级']
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    c.alignment = center
    c.border = thin

for idx, (cname, teacher, dow, start_node, step, st, et, room, weeks, credit) in enumerate(courses, 1):
    end_node = start_node + step - 1
    week_str = f'第{min(weeks)}-{max(weeks)}周'
    row_data = [idx, cname, teacher, day_names[dow-1], f'{start_node}-{end_node}节', st, et, room, week_str, credit, '人工智能2024级1班']
    for col, val in enumerate(row_data, 1):
        c = ws2.cell(row=idx+1, column=col, value=val)
        c.alignment = center
        c.border = thin
        if col == 2:  # 课程名列加底色
            c.fill = PatternFill(start_color=course_colors.get(cname, 'FFFFFF'), end_color=course_colors.get(cname, 'FFFFFF'), fill_type='solid')
            c.font = Font(bold=True)

widths2 = [6, 20, 10, 8, 10, 10, 10, 16, 12, 8, 20]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = 'A2'

# ============== Sheet3: 课程信息汇总 ==============
ws3 = wb.create_sheet('课程汇总')

ws3.merge_cells('A1:F1')
t3 = ws3['A1']
t3.value = '人工智能2024级1班 暑假培训课程汇总'
t3.font = Font(bold=True, size=14, color='FFFFFF')
t3.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
t3.alignment = center
ws3.row_dimensions[1].height = 30

headers3 = ['课程名称', '授课教师', '总课时', '周次', '上课时间', '教室']
for col, h in enumerate(headers3, 1):
    c = ws3.cell(row=2, column=col, value=h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    c.alignment = center
    c.border = thin

# 按课程汇总
course_summary = {}
for (cname, teacher, dow, start_node, step, st, et, room, weeks, credit) in courses:
    if cname not in course_summary:
        course_summary[cname] = {'teacher': teacher, 'credit': credit, 'weeks': set(), 'times': set(), 'rooms': set()}
    course_summary[cname]['weeks'].update(weeks)
    end_node = start_node + step - 1
    course_summary[cname]['times'].add(f'{day_names[dow-1]} {start_node}-{end_node}节({st}-{et})')
    course_summary[cname]['rooms'].add(room)

for idx, (cname, info) in enumerate(course_summary.items(), 1):
    week_list = sorted(info['weeks'])
    week_str = f'第{min(week_list)}-{max(week_list)}周'
    time_str = '\n'.join(sorted(info['times']))
    room_str = '、'.join(sorted(info['rooms']))
    row_data = [cname, info['teacher'], info['credit'], week_str, time_str, room_str]
    for col, val in enumerate(row_data, 1):
        c = ws3.cell(row=idx+2, column=col, value=val)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin
        if col == 1:
            c.fill = PatternFill(start_color=course_colors.get(cname, 'FFFFFF'), end_color=course_colors.get(cname, 'FFFFFF'), fill_type='solid')
            c.font = Font(bold=True)
    ws3.row_dimensions[idx+2].height = 50

widths3 = [20, 12, 8, 12, 30, 18]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

output_path = r'C:\Users\jay\Desktop\myBishe\aiStudy\init-db\summer-training-schedule.xlsx'
wb.save(output_path)
print(f'\nExcel已生成: {output_path}')
print(f'共 {len(courses)} 条排课记录，{len(course_summary)} 门课程')
for cname, info in course_summary.items():
    print(f'  - {cname}（{info["teacher"]}）: {info["credit"]}课时')
