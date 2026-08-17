# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '软件技术2024级1班课表'

# ===== 样式定义 =====
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
title_font = Font(size=16, bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 不同课程用不同颜色
color_map = {
    '高等数学': 'FFF2CC',
    'Java程序设计': 'DAEEF3',
    '大学英语': 'E2EFDA',
    '数据库原理': 'FCE4D6',
    '大学语文': 'FFE6F0',
    '体育': 'D9E1F2',
    '思政课': 'F2DCDB',
}

# ===== 标题 =====
ws.merge_cells('A1:H1')
ws['A1'] = '软件技术2024级1班 课表 (2025-2026-2学期)'
ws['A1'].font = title_font
ws['A1'].alignment = center_align
ws.row_dimensions[1].height = 32

# ===== 表头 =====
headers = ['星期', '节次', '开始时间', '结束时间', '课程名称', '授课教师', '教室', '周次']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# ===== 多样化课表数据 =====
# 包含：不同节次（1-12节）、单双周、不同教室、晚课、连续多节、分段周次
data = [
    # 周一：早 + 上午 + 下午
    ('周一', 1, '08:00', '08:45', '高等数学', '李建国', 'A101', '1-16周'),
    ('周一', 2, '08:55', '09:40', '高等数学', '李建国', 'A101', '1-16周'),
    ('周一', 3, '10:00', '10:45', 'Java程序设计', '陈志强', 'B201', '1-16周'),
    ('周一', 4, '10:55', '11:40', 'Java程序设计', '陈志强', 'B201', '1-16周'),
    ('周一', 7, '14:00', '14:45', '体育', '陈志强', '体育馆', '1-16周'),
    ('周一', 8, '14:55', '15:40', '体育', '陈志强', '体育馆', '1-16周'),
    ('周一', 9, '16:00', '16:45', '大学英语', '王美玲', 'A102', '单周'),
    ('周一', 10, '16:55', '17:40', '大学英语', '王美玲', 'A102', '单周'),

    # 周二：早 + 下午 + 晚
    ('周二', 1, '08:00', '08:45', '大学英语', '王美玲', 'A102', '1-16周'),
    ('周二', 2, '08:55', '09:40', '大学英语', '王美玲', 'A102', '1-16周'),
    ('周二', 5, '14:00', '14:45', '数据库原理', '刘晓燕', 'B202', '1-16周'),
    ('周二', 6, '14:55', '15:40', '数据库原理', '刘晓燕', 'B202', '1-16周'),
    ('周二', 11, '19:00', '19:45', '思政课', '郑国栋', 'C301', '双周'),
    ('周二', 12, '19:55', '20:40', '思政课', '郑国栋', 'C301', '双周'),

    # 周三：早 + 上午 + 下午
    ('周三', 1, '08:00', '08:45', '大学语文', '张雅琴', 'A103', '1-16周'),
    ('周三', 2, '08:55', '09:40', '大学语文', '张雅琴', 'A103', '1-16周'),
    ('周三', 3, '10:00', '10:45', '高等数学', '李建国', 'A101', '1-16周'),
    ('周三', 4, '10:55', '11:40', '高等数学', '李建国', 'A101', '1-16周'),
    ('周三', 7, '14:00', '14:45', 'Java程序设计', '陈志强', 'B201', '1-8周'),
    ('周三', 8, '14:55', '15:40', 'Java程序设计', '陈志强', 'B201', '1-8周'),

    # 周四：早 + 上午 + 下午
    ('周四', 1, '08:00', '08:45', '大学英语', '王美玲', 'A102', '1-16周'),
    ('周四', 2, '08:55', '09:40', '大学英语', '王美玲', 'A102', '1-16周'),
    ('周四', 3, '10:00', '10:45', 'Java程序设计', '陈志强', 'B201', '1-16周'),
    ('周四', 4, '10:55', '11:40', 'Java程序设计', '陈志强', 'B201', '1-16周'),
    ('周四', 7, '14:00', '14:45', '数据库原理', '刘晓燕', 'B202', '1-16周'),
    ('周四', 8, '14:55', '15:40', '数据库原理', '刘晓燕', 'B202', '1-16周'),

    # 周五：早 + 上午 + 下午 + 晚
    ('周五', 1, '08:00', '08:45', '大学语文', '张雅琴', 'A103', '1-16周'),
    ('周五', 2, '08:55', '09:40', '大学语文', '张雅琴', 'A103', '1-16周'),
    ('周五', 3, '10:00', '10:45', '数据库原理', '刘晓燕', 'B202', '1-16周'),
    ('周五', 4, '10:55', '11:40', '数据库原理', '刘晓燕', 'B202', '1-16周'),
    ('周五', 9, '16:00', '16:45', '高等数学', '李建国', 'A101', '9-16周'),
    ('周五', 10, '16:55', '17:40', '高等数学', '李建国', 'A101', '9-16周'),
    ('周五', 11, '19:00', '19:45', '大学语文', '张雅琴', 'A103', '单周'),
    ('周五', 12, '19:55', '20:40', '大学语文', '张雅琴', 'A103', '单周'),
]

# ===== 写入数据 =====
for row_idx, row_data in enumerate(data, 3):
    course = row_data[4]
    fill_color = color_map.get(course, 'FFFFFF')
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

# ===== 说明区 =====
note_row = len(data) + 4
ws.merge_cells(f'A{note_row}:H{note_row}')
ws.cell(row=note_row, column=1, value='备注：').font = Font(bold=True)
note_row += 1
notes = [
    '1. 节次说明：1-4节为上午，5-8节为下午，9-12节为晚间',
    '2. 周次说明：1-16周为全学期；单周为1,3,5,7,9,11,13,15周；双周为2,4,6,8,10,12,14,16周',
    '3. 9-16周表示课程从第9周开始，至第16周结束',
    '4. 同一课程可能在不同星期、不同节次出现，请按实际课表填写',
]
for i, note in enumerate(notes):
    ws.merge_cells(f'A{note_row+i}:H{note_row+i}')
    cell = ws.cell(row=note_row+i, column=1, value=note)
    cell.font = Font(size=10, color='666666')
    cell.alignment = Alignment(horizontal='left', vertical='center')

# ===== 列宽 =====
widths = [8, 8, 12, 12, 18, 12, 12, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 冻结表头
ws.freeze_panes = 'A3'

wb.save(r'C:\Users\jay\Desktop\myBishe\aiStudy\init-db\test-schedule-import.xlsx')
print(f'Excel saved with {len(data)} records')
