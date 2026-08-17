# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '学生导入模板'

# 样式
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center')

# 表头
headers = ['学号', '姓名', '用户名', '班级名称', '专业', '年级', '邮箱', '手机号']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin

# 示例数据（5条）
sample = [
    ('2024036', '赵小明', 'zhaoxm', '软件技术2024级1班', '软件技术', '2024', 'zhaoxm@example.com', '13900000036'),
    ('2024037', '钱小红', '', '软件技术2024级1班', '软件技术', '2024', '', ''),
    ('2024038', '孙小刚', 'sunxg', '', '软件技术', '2024', '', ''),
    ('2024039', '李丽', 'lili2024', '软件技术2024级1班', '', '', 'lili@example.com', '13900000039'),
    ('2024040', '周强', 'zhouq', '软件技术2024级1班', '软件技术', '2024', '', ''),
]

for row_idx, row_data in enumerate(sample, 2):
    for col_idx, val in enumerate(row_data, 1):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.alignment = center
        c.border = thin

# 列宽
widths = [12, 10, 12, 20, 12, 8, 22, 14]
from openpyxl.utils import get_column_letter
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 说明
note_row = len(sample) + 3
ws.cell(row=note_row, column=1, value='说明：').font = Font(bold=True)
notes = [
    '1. 第1行为表头，请勿修改；从第2行开始填写学生数据',
    '2. 学号、姓名为必填；用户名为空时系统自动用学号作为用户名',
    '3. 班级名称需与系统中已有班级完全一致（可留空，表示暂不分班）',
    '4. 专业、年级、邮箱、手机号为选填项',
    '5. 默认密码为 123456，导入后可单独重置',
    '6. 用户名已存在的行会自动跳过，不影响其他行导入',
]
for i, note in enumerate(notes):
    ws.cell(row=note_row + 1 + i, column=1, value=note).font = Font(size=10, color='666666')

wb.save(r'C:\Users\jay\Desktop\myBishe\aiStudy\init-db\student-import-template.xlsx')
print(f'Template saved with {len(sample)} sample rows')
