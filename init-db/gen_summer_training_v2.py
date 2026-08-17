# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 系统固定节次时间段（艺术学部/汽车学部作息）
NODE_TIMES = {
    1: ('08:10', '08:50'),
    2: ('09:00', '09:40'),
    3: ('09:50', '10:30'),
    4: ('10:40', '11:20'),
    5: ('15:10', '15:50'),
    6: ('16:00', '16:40'),
    7: ('19:50', '20:10'),
    8: ('20:20', '21:00'),
}

# 课程安排（已按系统 8 节课表校正）
# 字段说明：course_name, teacher, day_of_week, start_node, step, classroom, weeks, credit, semester
# credit 在本系统中用于排课弹窗的周课时上限，因此取每周实际占用节数
courses = [
    # Python：周一/周三/周五 上午 1-4 节
    ("Python编程入门", "李明远", 1, 1, 4, "实训楼A201", [1, 2, 3, 4, 5, 6], 12, "2026-暑假培训"),
    ("Python编程入门", "李明远", 3, 1, 4, "实训楼A201", [1, 2, 3, 4, 5, 6], 12, "2026-暑假培训"),
    ("Python编程入门", "李明远", 5, 1, 4, "实训楼A201", [1, 2, 3, 4, 5, 6], 12, "2026-暑假培训"),

    # 机器学习：周二/周四/周六 上午 1-4 节
    ("机器学习基础", "王思远", 2, 1, 4, "实训楼A202", [1, 2, 3, 4, 5, 6], 12, "2026-暑假培训"),
    ("机器学习基础", "王思远", 4, 1, 4, "实训楼A202", [1, 2, 3, 4, 5, 6], 12, "2026-暑假培训"),
    ("机器学习基础", "王思远", 6, 1, 4, "实训楼A202", [1, 2, 3, 4, 5, 6], 12, "2026-暑假培训"),

    # 深度学习：周一/周三/周五 下午/晚上 5-8 节
    ("深度学习实践", "赵晓峰", 1, 5, 4, "实训楼B301", [2, 3, 4, 5, 6], 12, "2026-暑假培训"),
    ("深度学习实践", "赵晓峰", 3, 5, 4, "实训楼B301", [2, 3, 4, 5, 6], 12, "2026-暑假培训"),
    ("深度学习实践", "赵晓峰", 5, 5, 4, "实训楼B301", [2, 3, 4, 5, 6], 12, "2026-暑假培训"),

    # 数据可视化：周二/周四/周六 下午/晚上 5-8 节
    ("数据可视化", "陈雨桐", 2, 5, 4, "实训楼B302", [2, 3, 4, 5, 6], 12, "2026-暑假培训"),
    ("数据可视化", "陈雨桐", 4, 5, 4, "实训楼B302", [2, 3, 4, 5, 6], 12, "2026-暑假培训"),
    ("数据可视化", "陈雨桐", 6, 5, 4, "实训楼B302", [2, 3, 4, 5, 6], 12, "2026-暑假培训"),

    # 自然语言处理：周二/周四 晚上 7-8 节（修正原 9-10 节超出系统范围的问题）
    ("自然语言处理入门", "孙文博", 2, 7, 2, "线上直播", [3, 4, 5, 6], 4, "2026-暑假培训"),
    ("自然语言处理入门", "孙文博", 4, 7, 2, "线上直播", [3, 4, 5, 6], 4, "2026-暑假培训"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '暑假培训课表'

# 标题
ws.merge_cells('A1:I1')
title_cell = ws['A1']
title_cell.value = '人工智能2024级1班 暑假培训课表（系统 8 节课表）'
title_cell.font = Font(bold=True, size=16, color='FFFFFF')
title_cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 36

# 副标题
ws.merge_cells('A2:I2')
sub_cell = ws['A2']
sub_cell.value = '培训时间：2026年7月25日 - 2026年8月31日（共6周） | 学期：2026-暑假培训'
sub_cell.font = Font(size=11, color='666666')
sub_cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
sub_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 24

day_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 表头
ws.cell(row=3, column=1, value='节次')
ws.cell(row=3, column=2, value='时间')
for i, day in enumerate(day_names):
    ws.cell(row=3, column=3 + i, value=day)

for col in range(1, 10):
    c = ws.cell(row=3, column=col)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin

# 构建可视化网格，课程跨多节时纵向合并
course_colors = {
    'Python编程入门': 'FFF2CC',
    '机器学习基础': 'E2EFDA',
    '深度学习实践': 'DDEBF7',
    '数据可视化': 'FCE4D6',
    '自然语言处理入门': 'E4DFEC'
}

schedule_grid = [[None] * 7 for _ in range(8)]

for (cname, teacher, dow, start_node, step, room, weeks, credit, semester) in courses:
    display = (cname, teacher, room, weeks, step)
    for offset in range(step):
        schedule_grid[start_node - 1 + offset][dow - 1] = display

row_num = 4
for node in range(1, 9):
    start_t, end_t = NODE_TIMES[node]
    ws.cell(row=row_num, column=1, value=f'第{node}节')
    ws.cell(row=row_num, column=2, value=f'{start_t}-{end_t}')

    for day_idx in range(7):
        cell_info = schedule_grid[node - 1][day_idx]
        if cell_info:
            cname, teacher, room, weeks, _ = cell_info
            # 只在连续课程的起始节显示内容，避免重复
            is_start = node == 1 or schedule_grid[node - 2][day_idx] != cell_info
            if is_start:
                display_text = f'{cname}\n@{teacher}\n{room}\n第{min(weeks)}-{max(weeks)}周'
                cell = ws.cell(row=row_num, column=3 + day_idx, value=display_text)
                cell.fill = PatternFill(start_color=course_colors[cname], end_color=course_colors[cname], fill_type='solid')
                cell.font = Font(size=9, bold=True)
                cell.alignment = center
                cell.border = thin
                # 计算需要合并的节数
                merge_end = node
                while merge_end < 8 and schedule_grid[merge_end][day_idx] == cell_info:
                    merge_end += 1
                if merge_end > node:
                    ws.merge_cells(start_row=row_num, start_column=3 + day_idx, end_row=row_num + merge_end - node, end_column=3 + day_idx)
        else:
            cell = ws.cell(row=row_num, column=3 + day_idx, value='')
            cell.border = thin
            cell.alignment = center

    for col in range(1, 10):
        ws.cell(row=row_num, column=col).border = thin
        ws.cell(row=row_num, column=col).alignment = center

    ws.row_dimensions[row_num].height = 36
    row_num += 1

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 14
for i in range(7):
    ws.column_dimensions[get_column_letter(3 + i)].width = 18

# 课程明细表（导入实际使用的数据源）
ws2 = wb.create_sheet('课程明细')
headers2 = ['序号', '课程名称', '授课教师', '星期', '节次', '开始时间', '结束时间', '教室', '周次', '总课时', '班级', '学期']
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = header_fill
    c.alignment = center
    c.border = thin

for idx, (cname, teacher, dow, start_node, step, room, weeks, credit, semester) in enumerate(courses, 1):
    end_node = start_node + step - 1
    start_t, _ = NODE_TIMES[start_node]
    _, end_t = NODE_TIMES[end_node]
    week_str = f'第{min(weeks)}-{max(weeks)}周'
    row_data = [idx, cname, teacher, day_names[dow - 1], f'{start_node}-{end_node}节', start_t, end_t, room, week_str, credit, '人工智能2024级1班', semester]
    for col, val in enumerate(row_data, 1):
        c = ws2.cell(row=idx + 1, column=col, value=val)
        c.alignment = center
        c.border = thin
        if col == 2:
            c.fill = PatternFill(start_color=course_colors[cname], end_color=course_colors[cname], fill_type='solid')
            c.font = Font(bold=True)

widths2 = [6, 20, 10, 8, 10, 10, 10, 16, 12, 8, 20, 16]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = 'A2'

output_path = r'C:\Users\jay\Desktop\myBishe\aiStudy\init-db\summer-training-schedule.xlsx'
wb.save(output_path)
print(f'已生成：{output_path}')
print(f'共 {len(courses)} 条课程明细记录')
