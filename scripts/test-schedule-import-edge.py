# -*- coding: utf-8 -*-
"""
课表导入边缘测试脚本

验证教师姓名容错逻辑：
1. 正确姓名 → 精确匹配（matched）
2. 错别字姓名 → 模糊匹配（fuzzy），并给出推荐教师
3. 不存在姓名 → 未匹配（unmatched），进入错误列表

用法（在仓库根目录执行）：
    python scripts/test-schedule-import-edge.py

环境变量：
    BASE_URL    后端地址，默认 http://localhost:8080
    ADMIN_USER  管理员账号，默认 admin
    ADMIN_PASS  管理员密码，默认 admin123
"""
import os
import sys
import tempfile
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE_URL = os.environ.get('BASE_URL', 'http://localhost:8080').rstrip('/')
ADMIN_USER = os.environ.get('ADMIN_USER', 'admin')
ADMIN_PASS = os.environ.get('ADMIN_PASS', 'admin123')


def login():
    url = f"{BASE_URL}/api/auth/login"
    resp = requests.post(url, json={"username": ADMIN_USER, "password": ADMIN_PASS}, timeout=10)
    if resp.status_code != 200:
        raise RuntimeError(f"登录失败: {resp.status_code} {resp.text}")
    data = resp.json()
    token = data.get("token") or data.get("data", {}).get("token")
    if not token:
        raise RuntimeError(f"登录响应缺少 token: {data}")
    return token


def get_teachers(token):
    url = f"{BASE_URL}/api/admin/user/teachers"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=10)
    if resp.status_code != 200:
        raise RuntimeError(f"获取教师列表失败: {resp.status_code} {resp.text}")
    return resp.json()


def build_test_excel(exact_teacher, class_name, semester):
    """构造包含三种教师姓名的测试课表。"""
    fuzzy_teacher = exact_teacher[:-1] + ("明" if exact_teacher[-1] != "明" else "民")
    missing_teacher = "系统中一定不存在的教师名XYZ"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "课程明细"

    headers = ['序号', '课程名称', '授课教师', '星期', '节次', '开始时间', '结束时间', '教室', '周次', '总课时', '班级', '学期']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin

    rows = [
        [1, '边缘测试-精确匹配', exact_teacher, '周一', '1-1节', '08:10', '08:50', 'A101', '[1,2]', 2, class_name, semester],
        [2, '边缘测试-错别字', fuzzy_teacher, '周一', '2-2节', '09:00', '09:40', 'A102', '[1,2]', 2, class_name, semester],
        [3, '边缘测试-不存在', missing_teacher, '周一', '3-3节', '09:50', '10:30', 'A103', '[1,2]', 2, class_name, semester],
    ]

    for r_idx, row_data in enumerate(rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.alignment = center
            c.border = thin

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14

    fd, path = tempfile.mkstemp(suffix='.xlsx', prefix='schedule_import_edge_test_')
    os.close(fd)
    wb.save(path)
    return path, fuzzy_teacher, missing_teacher


def preview_import(token, file_path):
    url = f"{BASE_URL}/api/schedule/import/preview"
    with open(file_path, 'rb') as f:
        files = {'file': ('edge-test.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        resp = requests.post(url, files=files, headers={"Authorization": f"Bearer {token}"}, timeout=60)
    if resp.status_code != 200:
        raise RuntimeError(f"预览接口调用失败: {resp.status_code} {resp.text}")
    return resp.json()


def confirm_import(token, preview_items):
    url = f"{BASE_URL}/api/schedule/import/confirm"
    # 对于 fuzzy 项，使用后端推荐的第一个建议教师（teacherId 已在前端选择逻辑中设置）
    payload = []
    for item in preview_items:
        payload.append(item)
    resp = requests.post(url, json=payload, headers={"Authorization": f"Bearer {token}"}, timeout=60)
    if resp.status_code != 200:
        raise RuntimeError(f"确认导入接口调用失败: {resp.status_code} {resp.text}")
    return resp.json()


def find_class(token):
    """获取一个可用班级。"""
    url = f"{BASE_URL}/api/admin/user/classes"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=10)
    if resp.status_code != 200:
        # 降级：使用固定班级
        return "人工智能2024级1班"
    classes = resp.json()
    if classes:
        return classes[0].get("className", "人工智能2024级1班")
    return "人工智能2024级1班"


def main():
    print(f"=== 课表导入边缘测试开始 ===")
    print(f"后端地址: {BASE_URL}")

    token = login()
    print(f"管理员登录成功")

    teachers = get_teachers(token)
    if not teachers:
        print("错误：系统中没有教师，无法执行边缘测试")
        sys.exit(1)

    exact_teacher = teachers[0]["realName"]
    print(f"选用系统中教师: {exact_teacher}")

    class_name = find_class(token)
    semester = "2026-边缘测试"

    file_path, fuzzy_teacher, missing_teacher = build_test_excel(exact_teacher, class_name, semester)
    print(f"测试文件已生成: {file_path}")

    try:
        preview = preview_import(token, file_path)
        print(f"预览结果: 总计 {preview.get('total')} 条, 成功 {preview.get('success')} 条, 异常 {len(preview.get('errors', []))} 条")

        # 断言 1：精确匹配
        matched_items = [p for p in preview.get('preview', []) if p.get('teacherMatchStatus') == 'matched']
        if len(matched_items) != 1:
            print(f"失败：精确匹配应返回 1 条，实际 {len(matched_items)} 条")
            sys.exit(1)
        if matched_items[0].get('teacherName') != exact_teacher:
            print(f"失败：精确匹配的教师姓名不一致")
            sys.exit(1)
        print("通过：精确匹配识别正确")

        # 断言 2：模糊匹配
        fuzzy_items = [p for p in preview.get('preview', []) if p.get('teacherMatchStatus') == 'fuzzy']
        if len(fuzzy_items) != 1:
            print(f"失败：模糊匹配应返回 1 条，实际 {len(fuzzy_items)} 条")
            sys.exit(1)
        suggestions = fuzzy_items[0].get('teacherSuggestions', [])
        if not any(s.get('teacherName') == exact_teacher for s in suggestions):
            print(f"失败：模糊匹配的建议列表中应包含 {exact_teacher}，实际 {suggestions}")
            sys.exit(1)
        print("通过：模糊匹配给出正确推荐")

        # 断言 3：未匹配进入错误列表
        error_msgs = []
        for err in preview.get('errors', []):
            error_msgs.extend(err.get('errors', []))
        if not any(missing_teacher in msg for msg in error_msgs):
            print(f"失败：不存在教师 {missing_teacher} 应出现在错误列表中，实际错误: {error_msgs}")
            sys.exit(1)
        print("通过：不存在教师被正确拦截")

        # 模拟前端修正：把 fuzzy 项的 teacherId 设置为推荐教师
        for item in preview.get('preview', []):
            if item.get('teacherMatchStatus') == 'fuzzy':
                best = item.get('teacherSuggestions', [])[0]
                item['teacherId'] = best['teacherId']
                item['teacherMatchStatus'] = 'matched'

        confirm = confirm_import(token, preview.get('preview', []))
        print(f"确认导入结果: 成功 {confirm.get('imported')} 条, 跳过 {confirm.get('skipped')} 条")

        if confirm.get('imported') != 2:
            print(f"失败：确认导入应成功 2 条，实际 {confirm.get('imported')} 条")
            sys.exit(1)
        if confirm.get('skipped') != 0:
            print(f"警告：存在跳过记录: {confirm.get('messages')}")

        print("=== 课表导入边缘测试全部通过 ===")
    finally:
        try:
            os.remove(file_path)
        except OSError:
            pass


if __name__ == '__main__':
    main()
